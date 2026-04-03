import pytest
import tempfile
from pathlib import Path
import pandas as pd
import openpyxl
from results_exporter import ResultsExporter
from statistical_analysis import SamplingInfo


def _make_sampling_info(applied=True, total=1_000_000, sampled=400_000):
    return SamplingInfo(applied=applied, total_rows=total, sampled_rows=sampled)


def test_excel_header_contains_sampling_info_when_applied(tmp_path):
    exporter = ResultsExporter(output_dir=str(tmp_path))
    info = _make_sampling_info(applied=True, total=1_000_000, sampled=400_000)

    df = pd.DataFrame({'HR': [1.2], 'p': [0.03]})
    cox_results = {'model1': {'summary': df}}

    path = exporter.export_cox_results(cox_results, sampling_info=info)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    # 첫 번째 행이 샘플링 정보 포함 여부 확인
    first_row_values = [ws.cell(1, c).value for c in range(1, 5)]
    assert any("400,000" in str(v) for v in first_row_values if v)


def test_excel_header_no_sampling_row_when_not_applied(tmp_path):
    exporter = ResultsExporter(output_dir=str(tmp_path))
    info = _make_sampling_info(applied=False, total=500, sampled=500)

    df = pd.DataFrame({'HR': [1.1], 'p': [0.05]})
    cox_results = {'model1': {'summary': df}}

    path = exporter.export_cox_results(cox_results, sampling_info=info)

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    first_row_values = [ws.cell(1, c).value for c in range(1, 5)]
    # 샘플링 정보 행 없어야 함
    assert not any("층화" in str(v) for v in first_row_values if v)
